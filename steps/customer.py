import openpyxl

import unittest
from behave import *
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from utility import config_reader, read_write_excel
from utility import config_reader


# Login:-------------------------------------------------------------------
@given(u'Open Chrome browser')
def browser(context):
    context.driver = webdriver.Chrome()

@given(u'Enter the Guru99 site url')
def guru_url(context):
    URL_guru = config_reader.Read_config("login-info", "guru_url")
    context.driver.get(URL_guru)
    context.driver.maximize_window()
    context.driver.implicitly_wait(100)
    time.sleep(3)

@given(u'Enter user Id and password with role "{role}"')
def user_pss(context, role):

    flag= False
    username, password = read_write_excel.getUserName_pass("common", role)
    if username.strip() == "" or password.strip() == "":
        flag= True

    if flag ==False:
        user_id=config_reader.Read_config("login-info", "uname_NAME")
        context.driver.find_element(By.NAME, 'uid').send_keys(username)

        user_passwrd=config_reader.Read_config("login-info", "password_NAME")
        context.driver.find_element(By.NAME, 'password').send_keys(password)

    else:
        unittest.TestCase.assertTrue(flag, "role does not match")

@then(u'Click on login button')
def login(context):
    user_login=config_reader.Read_config("login-info", "login_NAME")
    context.driver.find_element(By.NAME, 'btnLogin').click()
    time.sleep(3)

@then(u'Verify customer logged in successfully')
def cust_login(context):
    status = context.driver.find_element(By.XPATH, '/html/body/table/tbody/tr/td/table/tbody/tr[3]/td').is_displayed()
    assert status is True


# ---------------------------------------------------------------------------------------
@given(u'Click on the new customer tab')
def step_impl(context):
    # New_Customer:
    new_cust= config_reader.Read_config("locators-customer", "new_cust_XPATH")

    context.driver.find_element(By.XPATH, '/html/body/div[3]/div/ul/li[2]/a').click()
    time.sleep(10)

@given(u'Enter the Customer Name "{cname}"')
def step_impl(context, cname):
    context.Customer_Name= read_write_excel.getCustomer_details("customer")

    cus_name= config_reader.Read_config("locators-customer", "cname_NAME")
    context.driver.find_element(By.NAME, 'name').send_keys(context.Customer_Name)

@given(u'Select Gender')
def step_impl(context):

    gender= config_reader.Read_config("locators-customer", "gender_XPATH")

    context.driver.find_element(By.XPATH, '/html/body/table/tbody/tr/td/table/tbody/tr[5]/td[2]/input[2]').click()


@given(u'Select Date of Birth "{dob}"')
def step_impl(context, dob):
    context.Date_of_Birth= read_write_excel.getCustomer_details("customer")
    dob= config_reader.Read_config("locators-customer", "dob_NAME")
    context.driver.find_element(By.NAME, 'dob').send_keys(context.Date_of_Birth)


@when(u'Enter the Cust Address "{add}"')
def step_impl(context, add):
    context.Cust_Address = read_write_excel.getCustomer_details("customer")
    dob= config_reader.Read_config("locators-customer", "add_NAME")

    context.driver.find_element(By.NAME, 'addr').send_keys(context.Cust_Address)


@when(u'Enter the city "{city}"')
def step_impl(context, city):
    context.city = read_write_excel.getCustomer_details("customer")

    city= config_reader.Read_config("locators-customer", "city_NAME")
    context.driver.find_element(By.NAME, 'city').send_keys(context.city)
    time.sleep(2)

@when(u'Enter the state "{state}"')
def step_impl(context, state):
    context.state= read_write_excel.getCustomer_details("customer")

    state= config_reader.Read_config("locators-customer", "state_NAME")

    context.driver.find_element(By.NAME, 'state').send_keys(context.state)
    time.sleep(3)

@when(u'Enter the PIN "{pin}"')
def step_impl(context, pin):
    context.PIN= read_write_excel.getCustomer_details("customer")

    pin= config_reader.Read_config("locators-customer", "pin_NAME")
    context.driver.find_element(By.NAME, 'pinno').send_keys(context.PIN)
    time.sleep(3)

@then(u'Enter the mobile number "{mobile}"')
def step_impl(context, mobile):
    context.mobile_number= read_write_excel.getCustomer_details("customer")

    mob= config_reader.Read_config("locators-customer", "tel_NAME")
    context.driver.find_element(By.NAME, 'telephoneno').send_keys(context.mobile_number)
    time.sleep(3)

@then(u'Enter email-id "{emailid}"')
def step_impl(context, emailid):

    context.email_id= read_write_excel.getCustomer_details("customer")

    email= config_reader.Read_config("locators-customer", "email_NAME")
    context.driver.find_element(By.NAME, 'emailid').send_keys(context.email_id)
    time.sleep(3)

@then(u'Enter password "{password_2}"')
def step_impl(context, password_2):
    context.customer_password= read_write_excel.getCustomer_details("customer")

    pass2= config_reader.Read_config("locators-customer", "password_NAME")

    context.driver.find_element(By.NAME, 'password').send_keys(context.customer_password)

@then(u'Click on the submit button of new customer section')
def step_impl(context):
    sub1= config_reader.Read_config("locators-customer", "sub_but1_XPATH")
    context.driver.find_element(By.XPATH, '/html/body/table/tbody/tr/td/table/tbody/tr[14]/td[2]/input[1]').click()
    time.sleep(5)

@then(u'Verify customer registration')
def step_impl(context):
    reg= config_reader.Read_config("locators-customer", "reg_XPATH")

    reg=context.driver.find_element(By.XPATH,'//*[@id="customer"]/tbody/tr[1]/td/p').is_displayed()
    assert reg is True


@then(u'Capture the customer ID and store in the variable')
def step_impl(context):
    # customer_id :
    cust_id= config_reader.Read_config("locators-customer", "custID_XPATH")

    context.cus_id = context.driver.find_element(By.XPATH, '//*[@id="customer"]/tbody/tr[4]/td[2]').text
    time.sleep(5)

    # write guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName=workbook["customer"]

    c = 1
    r= 2
    for r in range (2,r+1):

        dict1 = {'customer_id': context.cus_id}
        # to take out the length of the list
        cnt = len(dict1)

        # to write the data in the cells
        sheetName.cell(row=r, column=c).value = context.cus_id

        # Logic for arranging the data (list data) in the table in row format
        for i in range(2, cnt + 2):
            sheetName.cell(row=i, column=1).value = dict['customer_id']

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")



# --------------------------------------------------------------------------------------------

@given(u'Click on the Edit customer tab of edit customer section')
def step_impl(context):

    # Edit_customer:
    edit_cus= config_reader.Read_config("locators-customer", "editCust_PARTIAL_LINK_TEXT")

    context.driver.find_element(By.PARTIAL_LINK_TEXT,'Edit Customer').click()
    time.sleep(5)

    # Read guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["customer"]

    for i in range(2, sheetName.max_row + 1):
        for j in range(1,2):
            # for not printing the None
            context.cusid= sheetName.cell(row=i, column=j).value
            return context.cusid

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")



    cusid= config_reader.Read_config("locators-customer", "cusid_NAME")

    context.driver.find_element(By.NAME, 'cusid').send_keys(context.cusid)
    time.sleep(3)


@then(u'Click on the submit button of edit section')
def step_impl(context):
    sub2= config_reader.Read_config("locators-customer", "sub2_NAME")
    context.driver.find_element(By.NAME, 'AccSubmit').click()
    time.sleep(5)

@when(u'For Editing enter the Address')
def step_impl(context):

    context.driver.find_element(By.NAME, 'addr').clear()
    context.driver.find_element(By.NAME, 'addr').send_keys('Civil lines')
    context.driver.find_element(By.NAME, 'pinno').clear()
    context.driver.find_element(By.NAME, 'telephoneno').clear()
    context.driver.find_element(By.NAME, 'pinno').send_keys('440034')
    context.driver.find_element(By.NAME, 'telephoneno').send_keys('8482653695')

    time.sleep(2)
    context.driver.execute_script("window.scrollBy(0,800)")                        #scroll the window screen


@then(u'Click on the submit button of new edit section')
def step_impl(context):
    sub3= config_reader.Read_config("locators-customer", "sub3_NAME")
    context.driver.find_element(By.NAME,'sub').click()
    time.sleep(3)

@then(u'Switch to alert')
def step_impl(context):
    alrt = context.driver.switch_to.alert
    time.sleep(5)
    alrt.accept()
    time.sleep(5)

#-----------------------------------------------------------------------------------------------
#
# @given(u'Click on the Delete customer tab of delete customer section')
# def step_impl(context):
#
#     #Delete_customer:
#     context.driver.find_element(By.PARTIAL_LINK_TEXT,'Delete Customer').click()
#
# @when(u'Enter the customer ID of delete customer section')
# def step_impl(context):
#     # Read Excel file:
#     f = open("/guru99_cust_id.txt", "r")
#     if f.mode == "r":
#         context.contents = f.read()
#
#     context.driver.find_element(By.NAME,'cusid').send_keys(context.contents)
#
# @then(u'Click on the submit button of delete customer section')
# def step_impl(context):
#     context.driver.find_element(By.NAME,'AccSubmit').click()
#     time.sleep(5)
#
#     alrt = context.driver.switch_to.alert
#     time.sleep(5)
#     alrt.accept()
#     time.sleep(5)











