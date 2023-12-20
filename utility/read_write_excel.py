import openpyxl
from utility import config_reader, read_write_excel

from utility.config_reader import Read_config

filepath = Read_config('login-info', 'xl_path')

# filepath = "D:\\practice_project\\guru99_data_xl.xlsx"
workbook = openpyxl.load_workbook(filepath)



def getUserName_pass(sheetname, role):
    sheet = workbook[sheetname]
    un = ""
    password = ""
    max_row = sheet.max_row
    max_col = sheet.max_column
    for i in range(2, max_row + 1):
        if sheet.cell(row=i, column=1).value == role:
            un = sheet.cell(row=i, column=2).value
            password = sheet.cell(row=i, column=3).value
            return un, password
    if un == "":
        return un, password


def getCustomer_details(sheetname):
    sheet = workbook[sheetname]
    customer_name = ""
    dob = ""
    add = ""
    city = ""
    state = ""
    pin = ""
    mobile_no = ""
    email_id = ""
    customer_pass = ""

    max_row = sheet.max_row
    max_col = sheet.max_column
    sr_no= 2

    for i in range(2, max_row + 1):
        if sheet.cell(row=i, column=i).value == sr_no:
            customer_name = sheet.cell(row=i, column=2).value
            dob= sheet.cell(row=i, column=3).value
            add= sheet.cell(row=i, column=4).value
            city= sheet.cell(row=i, column=5).value
            state= sheet.cell(row=i, column=6).value
            pin= sheet.cell(row=i, column=7).value
            mobile_no= sheet.cell(row=i, column=8).value
            email_id= sheet.cell(row=i, column=9).value
            customer_pass= sheet.cell(row=i, column=10).value
            return customer_name, dob, add, city, state, pin, mobile_no, email_id, customer_pass
    if customer_name == "":
        return customer_name








